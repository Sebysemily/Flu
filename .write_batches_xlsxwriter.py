from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import csv
import xlsxwriter

root = Path('/home/sebas/Work_proyects/FLU/flu_aviar')
batch_dir = root / 'data' / 'gisaid' / 'batches_24_f' / 'batches_24'
source_path = batch_dir / 'gisaid_batch_uploader_mira_batch_04.xlsx'
out_04 = batch_dir / 'gisaid_batch_uploader_mira_batch_04.xlsx'
out_05 = batch_dir / 'gisaid_batch_uploader_mira_batch_05.xlsx'
summary_path = root / 'data' / 'input' / 'H5N1_EC_summary.csv'
fasta_path = root / 'data' / 'assembled' / 'ecuador_intermediate_sequences.fasta'
base_db = batch_dir / 'Base-datos-Flu-01.xlsx'

missing = ['Flu-0583','Flu-0584','Flu-0586','Flu-0589','Flu-0592','Flu-0593','Flu-0596','Flu-0599','Flu-0608','Flu-0611','Flu-0613','Flu-0614','Flu-0615','Flu-0619','Flu-0621','Flu-0622','Flu-0623','Flu-0641','Flu-0652','Flu-0654','Flu-0658']
fill_04 = missing[:11]
fill_05 = missing[11:]
segment_order = ['PB2', 'PB1', 'PA', 'HA', 'NP', 'NA', 'MP', 'NS']
authors_agro = 'Mateo Carvajal, Erika B. Muñoz, Denisse Benítez, Paul Leon, Roger Calvas, Rommel Guevara, David Jarrín, Euclides José De La Torre, Alexandra Paola Revelo, María Elena Rovalino, Verónica Barragán, Patricio Rojas-Silva, Gabriel Trueba, Michelle Grunauer, Paúl Cárdenas'
headers = ['Isolate_Id','Segment_Ids','Isolate_Name','Subtype','Lineage','Passage_History','Location','province','sub_province','Location_Additional_info','Host','Host_Additional_info','Seq_Id (HA)','Seq_Id (NA)','Seq_Id (PB1)','Seq_Id (PB2)','Seq_Id (PA)','Seq_Id (MP)','Seq_Id (NS)','Seq_Id (NP)','Seq_Id (HE)','Seq_Id (P3)','Submitting_Sample_Id','Authors','Originating_Lab_Id','Originating_Sample_Id','Collection_Month','Collection_Year','Collection_Date','Antigen_Character','Adamantanes_Resistance_geno','Oseltamivir_Resistance_geno','Zanamivir_Resistance_geno','Peramivir_Resistance_geno','Other_Resistance_geno','Adamantanes_Resistance_pheno','Oseltamivir_Resistance_pheno','Zanamivir_Resistance_pheno','Peramivir_Resistance_pheno','Other_Resistance_pheno','Host_Age','Host_Age_Unit','Host_Gender','Health_Status','Note','PMID','Publishing_Embargo']


def load_summary():
    summary = {}
    with summary_path.open('r', encoding='utf-8', newline='') as f:
        for row in csv.DictReader(f):
            if row['sample'] in missing and row['sample'] not in summary:
                summary[row['sample']] = row
    return summary


def load_origin():
    origin = {}
    wb_db = load_workbook(base_db, read_only=True, data_only=True)
    ws_db = wb_db['Listado']
    for r in range(2, ws_db.max_row + 1):
        code = ws_db.cell(r, 1).value
        if code in missing:
            origin[code] = ws_db.cell(r, 2).value
    wb_db.close()
    return origin


def load_seqs():
    seqs = {}
    current_id = None
    current_seq = []
    with fasta_path.open('r', encoding='utf-8') as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            if line.startswith('>'):
                if current_id is not None:
                    sample, segment = current_id.split('|', 1)
                    seqs.setdefault(sample, {})[segment.replace('A_', '')] = ''.join(current_seq)
                current_id = line[1:]
                current_seq = []
            else:
                current_seq.append(line)
    if current_id is not None:
        sample, segment = current_id.split('|', 1)
        seqs.setdefault(sample, {})[segment.replace('A_', '')] = ''.join(current_seq)
    return seqs


def build_row(sample, summary, origin, seqs):
    meta = summary[sample]
    dt = datetime.strptime(meta['date'], '%Y-%m-%d')
    available = [seg for seg in segment_order if seg in seqs.get(sample, {})]
    data = {
        'Isolate_Id': sample,
        'Segment_Ids': ','.join(available),
        'Isolate_Name': f'A/chicken/Ecuador/{sample}/{dt.year}',
        'Subtype': 'H5N1',
        'Lineage': None,
        'Passage_History': 'Original',
        'Location': 'Ecuador',
        'province': meta['province'].title() if meta.get('province') else None,
        'sub_province': None,
        'Location_Additional_info': None,
        'Host': 'Gallus gallus domesticus',
        'Host_Additional_info': 'Aves de corral',
        'Seq_Id (HA)': f'{sample}|A_HA' if 'HA' in available else None,
        'Seq_Id (NA)': f'{sample}|A_NA' if 'NA' in available else None,
        'Seq_Id (PB1)': f'{sample}|A_PB1' if 'PB1' in available else None,
        'Seq_Id (PB2)': f'{sample}|A_PB2' if 'PB2' in available else None,
        'Seq_Id (PA)': f'{sample}|A_PA' if 'PA' in available else None,
        'Seq_Id (MP)': f'{sample}|A_MP' if 'MP' in available else None,
        'Seq_Id (NS)': f'{sample}|A_NS' if 'NS' in available else None,
        'Seq_Id (NP)': f'{sample}|A_NP' if 'NP' in available else None,
        'Seq_Id (HE)': None,
        'Seq_Id (P3)': None,
        'Submitting_Sample_Id': sample,
        'Authors': authors_agro,
        'Originating_Lab_Id': 3536,
        'Originating_Sample_Id': origin.get(sample),
        'Collection_Month': f'{dt.month:02d}',
        'Collection_Year': dt.year,
        'Collection_Date': dt,
        'Antigen_Character': None,
        'Adamantanes_Resistance_geno': None,
        'Oseltamivir_Resistance_geno': None,
        'Zanamivir_Resistance_geno': None,
        'Peramivir_Resistance_geno': None,
        'Other_Resistance_geno': None,
        'Adamantanes_Resistance_pheno': None,
        'Oseltamivir_Resistance_pheno': None,
        'Zanamivir_Resistance_pheno': None,
        'Peramivir_Resistance_pheno': None,
        'Other_Resistance_pheno': None,
        'Host_Age': None,
        'Host_Age_Unit': None,
        'Host_Gender': None,
        'Health_Status': None,
        'Note': None,
        'PMID': None,
        'Publishing_Embargo': None,
    }
    return [data.get(h) for h in headers], available


def write_batch(path, include_source_rows, append_samples):
    summary = load_summary()
    origin = load_origin()
    seqs = load_seqs()
    src = load_workbook(source_path, read_only=True, data_only=False)
    wb = xlsxwriter.Workbook(str(path), {'constant_memory': True})

    # Isolates
    ws = wb.add_worksheet('Isolates')
    src_iso = src['Isolates']
    row_idx = 0
    header_row = next(src_iso.iter_rows(values_only=True))
    for col, value in enumerate(header_row):
        ws.write(row_idx, col, value)
    row_idx += 1
    if include_source_rows:
        for row in src_iso.iter_rows(values_only=True):
            pass
    for row in src_iso.iter_rows(values_only=True):
        if row == header_row:
            continue
        if include_source_rows:
            for col, value in enumerate(row):
                ws.write(row_idx, col, value)
            row_idx += 1
    for sample in append_samples:
        row_values, _ = build_row(sample, summary, origin, seqs)
        for col, value in enumerate(row_values):
            ws.write(row_idx, col, value)
        row_idx += 1

    # __shadow__
    ws = wb.add_worksheet('__shadow__')
    src_shadow = src['__shadow__']
    shadow_rows = list(src_shadow.iter_rows(values_only=True))
    row_idx = 0
    for col, value in enumerate(shadow_rows[0]):
        ws.write(row_idx, col, value)
    row_idx += 1
    if include_source_rows:
        for row in shadow_rows[1:]:
            for col, value in enumerate(row):
                ws.write(row_idx, col, value)
            row_idx += 1
    for sample in append_samples:
        _, segs = build_row(sample, summary, origin, seqs)
        ws.write(row_idx, 0, sample)
        ws.write(row_idx, 1, ','.join(segs))
        row_idx += 1

    # Sequences
    ws = wb.add_worksheet('Sequences')
    src_seq = src['Sequences']
    row_idx = 0
    if include_source_rows:
        for row in src_seq.iter_rows(values_only=True):
            for col, value in enumerate(row):
                ws.write(row_idx, col, value)
            row_idx += 1
    for sample in append_samples:
        _, segs = build_row(sample, summary, origin, seqs)
        for seg in segs:
            ws.write(row_idx, 0, f'>{sample}|A_{seg}')
            row_idx += 1
            ws.write(row_idx, 0, seqs[sample][seg])
            row_idx += 1

    # auxiliary sheets, copied as values only to preserve workbook shape
    for sname in src.sheetnames:
        if sname in {'Isolates', '__shadow__', 'Sequences'}:
            continue
        src_ws = src[sname]
        ws = wb.add_worksheet(sname)
        for r, row in enumerate(src_ws.iter_rows(values_only=True)):
            for c, value in enumerate(row):
                ws.write(r, c, value)

    src.close()
    wb.close()


write_batch(out_04, True, fill_04)
write_batch(out_05, False, fill_05)
print('done', len(fill_04), len(fill_05))
