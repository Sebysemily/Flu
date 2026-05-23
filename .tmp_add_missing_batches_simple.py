from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import csv
import shutil

root = Path('/home/sebas/Work_proyects/FLU/flu_aviar')
batch_dir = root / 'data' / 'gisaid' / 'batches_24_f' / 'batches_24'
base_batch = batch_dir / 'gisaid_batch_uploader_mira_batch_04.xlsx'
new_batch = batch_dir / 'gisaid_batch_uploader_mira_batch_05.xlsx'
summary_path = root / 'data' / 'input' / 'H5N1_EC_summary.csv'
fasta_path = root / 'data' / 'assembled' / 'ecuador_intermediate_sequences.fasta'
base_db = batch_dir / 'Base-datos-Flu-01.xlsx'

missing = ['Flu-0583','Flu-0584','Flu-0586','Flu-0589','Flu-0592','Flu-0593','Flu-0596','Flu-0599','Flu-0608','Flu-0611','Flu-0613','Flu-0614','Flu-0615','Flu-0619','Flu-0621','Flu-0622','Flu-0623','Flu-0641','Flu-0652','Flu-0654','Flu-0658']
fill_04 = missing[:11]
fill_05 = missing[11:]
segment_order = ['PB2', 'PB1', 'PA', 'HA', 'NP', 'NA', 'MP', 'NS']
authors_agro = 'Mateo Carvajal, Erika B. Muñoz, Denisse Benítez, Paul Leon, Roger Calvas, Rommel Guevara, David Jarrín, Euclides José De La Torre, Alexandra Paola Revelo, María Elena Rovalino, Verónica Barragán, Patricio Rojas-Silva, Gabriel Trueba, Michelle Grunauer, Paúl Cárdenas'

summary = {}
with summary_path.open('r', encoding='utf-8', newline='') as f:
    for row in csv.DictReader(f):
        if row['sample'] in missing and row['sample'] not in summary:
            summary[row['sample']] = row

origin = {}
wb_db = load_workbook(base_db, read_only=True, data_only=True)
ws_db = wb_db['Listado']
for r in range(2, ws_db.max_row + 1):
    code = ws_db.cell(r, 1).value
    if code in missing:
        origin[code] = ws_db.cell(r, 2).value
wb_db.close()

seqs = {}
current_id = None
current_seq = []
with fasta_path.open('r', encoding='utf-8') as f:
    for raw in f:
        line = raw.strip()
        if not line:
            continue
        if line.startswith('>'):
            if current_id is not None:
                sample, segment = current_id.split('|', 1)
                seqs.setdefault(sample, {})[segment.replace('A_', '')] = ''.join(current_seq)
            current_id = line[1:]
            current_seq = []
        else:
            current_seq.append(line)
    if current_id is not None:
        sample, segment = current_id.split('|', 1)
        seqs.setdefault(sample, {})[segment.replace('A_', '')] = ''.join(current_seq)

wb = load_workbook(base_batch)
iso = wb['Isolates']
shadow = wb['__shadow__']
seqsheet = wb['Sequences']
headers = [iso.cell(1, c).value for c in range(1, iso.max_column + 1)]


def build_row(sample):
    meta = summary[sample]
    dt = datetime.strptime(meta['date'], '%Y-%m-%d')
    available = [seg for seg in segment_order if seg in seqs.get(sample, {})]
    data = {
        'Isolate_Id': sample,
        'Segment_Ids': ','.join(available),
        'Isolate_Name': f'A/chicken/Ecuador/{sample}/{dt.year}',
        'Subtype': 'H5N1',
        'Lineage': None,
        'Passage_History': 'Original',
        'Location': 'Ecuador',
        'province': meta['province'].title() if meta.get('province') else None,
        'sub_province': None,
        'Location_Additional_info': None,
        'Host': 'Gallus gallus domesticus',
        'Host_Additional_info': 'Aves de corral',
        'Seq_Id (HA)': f'{sample}|A_HA' if 'HA' in available else None,
        'Seq_Id (NA)': f'{sample}|A_NA' if 'NA' in available else None,
        'Seq_Id (PB1)': f'{sample}|A_PB1' if 'PB1' in available else None,
        'Seq_Id (PB2)': f'{sample}|A_PB2' if 'PB2' in available else None,
        'Seq_Id (PA)': f'{sample}|A_PA' if 'PA' in available else None,
        'Seq_Id (MP)': f'{sample}|A_MP' if 'MP' in available else None,
        'Seq_Id (NS)': f'{sample}|A_NS' if 'NS' in available else None,
        'Seq_Id (NP)': f'{sample}|A_NP' if 'NP' in available else None,
        'Seq_Id (HE)': None,
        'Seq_Id (P3)': None,
        'Submitting_Sample_Id': sample,
        'Authors': authors_agro,
        'Originating_Lab_Id': 3536,
        'Originating_Sample_Id': origin.get(sample),
        'Collection_Month': f'{dt.month:02d}',
        'Collection_Year': dt.year,
        'Collection_Date': dt,
        'Antigen_Character': None,
        'Adamantanes_Resistance_geno': None,
        'Oseltamivir_Resistance_geno': None,
        'Zanamivir_Resistance_geno': None,
        'Peramivir_Resistance_geno': None,
        'Other_Resistance_geno': None,
        'Adamantanes_Resistance_pheno': None,
        'Oseltamivir_Resistance_pheno': None,
        'Zanamivir_Resistance_pheno': None,
        'Peramivir_Resistance_pheno': None,
        'Other_Resistance_pheno': None,
        'Host_Age': None,
        'Host_Age_Unit': None,
        'Host_Gender': None,
        'Health_Status': None,
        'Note': None,
        'PMID': None,
        'Publishing_Embargo': None,
    }
    return [data.get(h) for h in headers], available


def append_sample(ws_iso, ws_shadow, ws_seq, sample):
    row_values, segments = build_row(sample)
    ws_iso.append(row_values)
    ws_shadow.append([sample, ','.join(segments)])
    for seg in segments:
        ws_seq.append([f'>{sample}|A_{seg}'])
        ws_seq.append([seqs[sample][seg]])

for sample in fill_04:
    append_sample(iso, shadow, seqsheet, sample)
wb.save(base_batch)
wb.close()

shutil.copy2(base_batch, new_batch)
wb = load_workbook(new_batch)
iso = wb['Isolates']
shadow = wb['__shadow__']
seqsheet = wb['Sequences']
if iso.max_row > 1:
    iso.delete_rows(2, iso.max_row - 1)
if shadow.max_row > 1:
    shadow.delete_rows(2, shadow.max_row - 1)
if seqsheet.max_row > 1:
    seqsheet.delete_rows(2, seqsheet.max_row - 1)
for sample in fill_05:
    append_sample(iso, shadow, seqsheet, sample)
wb.save(new_batch)
wb.close()

print('batch04', len(fill_04))
print('batch05', len(fill_05))
print('done')
